import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


@dataclass(frozen=True)
class ColumnSpec:
    key: str
    group_name: str
    group_slug: str
    task_name: str
    task_slug: str
    sample_count: int | None
    excel_col: int


def slugify(value: str) -> str:
    raw = str(value or "").strip().lower()
    raw = re.sub(r"[^a-z0-9]+", "_", raw)
    raw = re.sub(r"_+", "_", raw).strip("_")
    return raw or "col"


def normalize_number(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in {"\\", "-", "—"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_int(value) -> int | None:
    num = normalize_number(value)
    if num is None:
        return None
    return int(round(num))


def normalize_date(value, wb_epoch) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        try:
            dt = from_excel(value, wb_epoch)
            if isinstance(dt, datetime):
                return dt.date().isoformat()
        except Exception:
            return None
    text = str(value).strip()
    return text or None


def find_col(ws, row: int, header: str) -> int | None:
    target = header.strip().lower()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=c).value
        if v is None:
            continue
        if str(v).strip().lower() == target:
            return c
    return None


def build_columns(ws, first_score_col: int, avg_col: int) -> list[ColumnSpec]:
    groups = []
    current = ""
    for c in range(first_score_col, avg_col):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip():
            current = str(v).strip()
        groups.append(current)

    seen_keys: dict[str, int] = {}
    columns: list[ColumnSpec] = []
    for idx, c in enumerate(range(first_score_col, avg_col)):
        group_name = groups[idx] or "Other"
        group_slug = slugify(group_name)
        task_name = ws.cell(row=2, column=c).value
        task_name = str(task_name).strip() if task_name is not None else f"Task {c}"
        task_slug = slugify(task_name)
        base_key = f"{group_slug}__{task_slug}"
        seen_keys[base_key] = seen_keys.get(base_key, 0) + 1
        key = base_key if seen_keys[base_key] == 1 else f"{base_key}__{seen_keys[base_key]}"

        columns.append(
            ColumnSpec(
                key=key,
                group_name=group_name,
                group_slug=group_slug,
                task_name=task_name,
                task_slug=task_slug,
                sample_count=normalize_int(ws.cell(row=3, column=c).value),
                excel_col=c,
            )
        )
    return columns


def main() -> None:
    parser = argparse.ArgumentParser(description="Export data/leaderboard.xlsx to a JSON file for the web leaderboard.")
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path("data/pairwise_acc.xlsx"),
        help="Path to leaderboard .xlsx (default: data/pairwise_acc.xlsx).",
    )
    parser.add_argument("--sheet", type=str, default="", help="Sheet name (default: first sheet).")
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("data/pairwise_acc.json"),
        help="Output path for JSON (default: data/pairwise_acc.json).",
    )
    args = parser.parse_args()

    if not args.xlsx.is_file():
        legacy = Path("data/leaderboard.xlsx")
        if legacy.is_file():
            args.xlsx = legacy
        else:
            raise SystemExit(f"--xlsx not found: {args.xlsx}")

    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    model_col = find_col(ws, 1, "Models") or 2
    param_col = find_col(ws, 1, "#Param (B)") or 3
    release_col = find_col(ws, 1, "Release time") or 4
    avg_col = find_col(ws, 1, "Average") or ws.max_column
    first_score_col = release_col + 1

    columns = build_columns(ws, first_score_col=first_score_col, avg_col=avg_col)

    models = []
    current_group = ""
    for r in range(4, ws.max_row + 1):
        group_cell = ws.cell(row=r, column=1).value
        if group_cell is not None and str(group_cell).strip():
            current_group = str(group_cell).strip()

        model_name = ws.cell(row=r, column=model_col).value
        if model_name is None or not str(model_name).strip():
            score_any = any(ws.cell(row=r, column=col.excel_col).value is not None for col in columns)
            avg_any = ws.cell(row=r, column=avg_col).value is not None
            if not score_any and not avg_any:
                continue
            model_name = ""
        model_name = str(model_name).strip()

        params_b = normalize_number(ws.cell(row=r, column=param_col).value)
        release_date = normalize_date(ws.cell(row=r, column=release_col).value, wb.epoch)
        average = normalize_number(ws.cell(row=r, column=avg_col).value) or 0.0

        scores: dict[str, float | None] = {}
        for col in columns:
            scores[col.key] = normalize_number(ws.cell(row=r, column=col.excel_col).value)

        models.append(
            {
                "group": current_group or None,
                "model": model_name,
                "paramsB": params_b,
                "releaseDate": release_date,
                "average": average,
                "scores": scores,
            }
        )

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "source": str(args.xlsx.as_posix()),
        "sheet": ws.title,
        "columns": [
            {
                "key": c.key,
                "groupName": c.group_name,
                "groupSlug": c.group_slug,
                "taskName": c.task_name,
                "taskSlug": c.task_slug,
                "sampleCount": c.sample_count,
            }
            for c in columns
        ],
        "models": models,
    }

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
